import io
from datetime import datetime, date, time
from functools import wraps

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import (
    Blueprint, render_template, request,
    jsonify, redirect, url_for, send_file, flash
)
from flask_login import login_required, current_user

from app.extensions import db
from app.models.contrato import Contrato
from app.models.reporte_operacional import ReporteOperacional
from app.models.user import User
from app.models.persona import Persona
from app.models.meta_operativa import MetaOperativa
from app.models.tipo_desvio import TipoDesvio
from app.models.parametro_neo import ParametroNeo
from app.models.actividad import Actividad
from app.models.accion_tomar import AccionTomar
from app.models.parametro_coor import ParametroCoor
from app.models.recurso_contrato import RecursoContrato
from app.models.placa_contrato import PlacaContrato
from app.models.user_contrato import UserContrato
from app.models.supervisor import Supervisor
from app.models.he_config import HeConfig
from app.models.hora_extra import HoraExtra, CONCEPTOS_HE
from app.models.he_corte import HeCorte


admin_bp = Blueprint("admin_bp", __name__, url_prefix="/admin")


def _sync_nombres_desde_personas():
    """Rellena HoraExtra.nombre donde está vacío usando la tabla personas."""
    personas_map = {p.Documento: p.Nombre for p in Persona.query.all()}
    registros = HoraExtra.query.filter(
        db.or_(HoraExtra.nombre == None, HoraExtra.nombre == "")
    ).all()
    actualizados = 0
    for he in registros:
        nombre_nuevo = personas_map.get(he.cedula)
        if nombre_nuevo:
            he.nombre = nombre_nuevo
            actualizados += 1
    if actualizados:
        db.session.commit()
    return actualizados


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.rol.lower() != "admin":
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


# ============================================================
# DASHBOARD
# ============================================================

@admin_bp.route("/")
@admin_required
def dashboard():
    stats = {
        "contratos":    Contrato.query.count(),
        "usuarios":     User.query.filter_by(activo=True).count(),
        "personas":     Persona.query.count(),
        "metas":        MetaOperativa.query.count(),
        "desvios":      TipoDesvio.query.count(),
        "parametros":   ParametroNeo.query.count(),
        "actividades":  Actividad.query.count(),
        "recursos_neo": RecursoContrato.query.filter(
            db.or_(
                RecursoContrato.recurso.like("Centro Técnico%"),
                RecursoContrato.recurso.like("SEDE %")
            )
        ).count(),
    }
    return render_template("admin/dashboard.html", stats=stats)


# ============================================================
# CONTRATOS
# ============================================================

@admin_bp.route("/contratos")
@admin_required
def contratos():
    lista = Contrato.query.order_by(Contrato.contrato).all()
    todos_usuarios = User.query.filter(
        User.rol.in_(["coordinador", "director"])
    ).order_by(User.nombre_completo).all()
    # Para cada contrato: qué usuarios tienen ese contrato asignado en UserContrato
    uc_all = UserContrato.query.all()
    usuarios_por_contrato = {}
    for c in lista:
        usuarios_por_contrato[c.contrato] = [
            u for u in todos_usuarios
            if any(uc.user_id == u.id and uc.contrato == c.contrato for uc in uc_all)
        ]
    return render_template("admin/contratos.html",
                           contratos=lista,
                           usuarios_por_contrato=usuarios_por_contrato)


@admin_bp.route("/api/contratos", methods=["POST"])
@admin_required
def api_crear_contrato():
    d = request.get_json() or {}
    nombre_completo = (d.get("contrato") or "").strip()
    if not nombre_completo:
        return jsonify({"success": False, "mensaje": "El nombre completo del contrato es requerido"}), 400
    if Contrato.query.filter_by(contrato=nombre_completo).first():
        return jsonify({"success": False, "mensaje": "Ya existe un contrato con ese nombre"}), 400
    codigo = (d.get("codigo") or "").strip() or None
    if codigo and Contrato.query.filter_by(codigo=codigo).first():
        return jsonify({"success": False, "mensaje": "Ya existe un contrato con ese código"}), 400
    c = Contrato(
        contrato=nombre_completo,
        codigo=codigo,
        nombre=(d.get("nombre") or "").strip() or None,
        sede=(d.get("sede") or "").strip() or None,
        proceso=(d.get("proceso") or "").strip() or None,
        activo=True
    )
    db.session.add(c)
    db.session.commit()
    return jsonify({"success": True, "id": c.id})


@admin_bp.route("/api/contratos/<int:id>", methods=["PUT"])
@admin_required
def api_editar_contrato(id):
    c = db.session.get(Contrato, id)
    if not c:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    if d.get("contrato"):
        c.contrato = d["contrato"].strip()
    if "codigo" in d:
        nuevo_cod = (d["codigo"] or "").strip() or None
        if nuevo_cod and nuevo_cod != c.codigo:
            if Contrato.query.filter_by(codigo=nuevo_cod).first():
                return jsonify({"success": False, "mensaje": "Ya existe ese código"}), 400
        c.codigo = nuevo_cod
    if "nombre" in d:
        c.nombre = (d["nombre"] or "").strip() or None
    if "sede" in d:
        c.sede = (d["sede"] or "").strip() or None
    if "proceso" in d:
        c.proceso = (d["proceso"] or "").strip() or None
    if "activo" in d:
        c.activo = bool(d["activo"])
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/contratos/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_contrato(id):
    c = db.session.get(Contrato, id)
    if not c:
        return jsonify({"success": False}), 404
    db.session.delete(c)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# USUARIOS
# ============================================================

@admin_bp.route("/usuarios")
@admin_required
def usuarios():
    lista = User.query.filter(
        ~User.rol.in_(["neo", "admin"])
    ).order_by(User.rol, User.nombre_completo).all()
    contratos = Contrato.query.order_by(Contrato.contrato).all()

    asignados_por_usuario = {}
    for u in lista:
        asignados_por_usuario[u.id] = {
            uc.contrato for uc in
            UserContrato.query.filter_by(user_id=u.id).all()
        }

    roles = ["coordinador", "director", "parqueadero", "admin_parqueadero"]
    return render_template("admin/usuarios.html",
                           usuarios=lista, contratos=contratos,
                           asignados_por_usuario=asignados_por_usuario,
                           roles=roles)


@admin_bp.route("/neo-usuarios")
@admin_required
def neo_usuarios():
    lista = User.query.filter(
        User.rol.in_(["neo", "admin"])
    ).order_by(User.rol, User.nombre_completo).all()
    contratos = Contrato.query.order_by(Contrato.contrato).all()

    # Pre-cargar contratos asignados de todos los usuarios de la lista
    asignados_por_usuario = {}
    for u in lista:
        asignados_por_usuario[u.id] = {
            uc.contrato for uc in
            UserContrato.query.filter_by(user_id=u.id).all()
        }

    return render_template(
        "admin/neo_usuarios.html",
        usuarios=lista,
        contratos=contratos,
        asignados_por_usuario=asignados_por_usuario
    )


@admin_bp.route("/parqueadero-usuarios")
@admin_required
def parqueadero_usuarios():
    lista = User.query.filter(
        User.rol.in_(["parqueadero", "admin_parqueadero"])
    ).order_by(User.rol, User.nombre_completo).all()
    return render_template("admin/parqueadero_usuarios.html", usuarios=lista)


@admin_bp.route("/api/usuarios", methods=["POST"])
@admin_required
def api_crear_usuario():
    d = request.get_json() or {}
    username = (d.get("username") or "").strip()
    nombre   = (d.get("nombre_completo") or "").strip()
    password = (d.get("password") or "").strip()
    rol      = (d.get("rol") or "").strip()
    if not all([username, nombre, password, rol]):
        return jsonify({"success": False, "mensaje": "Usuario, nombre, contraseña y rol son requeridos"}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({"success": False, "mensaje": "El usuario ya existe"}), 400
    u = User(
        username=username,
        nombre_completo=nombre,
        password_hash=password,
        rol=rol,
        contrato=d.get("contrato") or None,
        activo=True
    )
    db.session.add(u)
    db.session.commit()
    return jsonify({"success": True, "id": u.id})


@admin_bp.route("/api/usuarios/<int:id>", methods=["PUT"])
@admin_required
def api_editar_usuario(id):
    u = db.session.get(User, id)
    if not u:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    u.nombre_completo = (d.get("nombre_completo") or u.nombre_completo).strip()
    u.rol      = (d.get("rol") or u.rol).strip()
    u.contrato = d.get("contrato") or None
    u.activo   = bool(d.get("activo", u.activo))
    u.acceso_dashboard = bool(d.get("acceso_dashboard", u.acceso_dashboard))
    if "permiso" in d:
        u.set_permiso(d["permiso"], bool(d.get("valor", False)))
    if d.get("password"):
        u.password_hash = d["password"].strip()
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/usuarios/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_usuario(id):
    u = db.session.get(User, id)
    if not u:
        return jsonify({"success": False}), 404
    db.session.delete(u)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# PERSONAS
# ============================================================

@admin_bp.route("/personas")
@admin_required
def personas():
    lista = Persona.query.order_by(Persona.Nombre).all()
    return render_template("admin/personas.html", personas=lista)


@admin_bp.route("/api/personas", methods=["POST"])
@admin_required
def api_crear_persona():
    d = request.get_json() or {}
    doc    = (d.get("Documento") or "").strip()
    nombre = (d.get("Nombre") or "").strip()
    cargo  = (d.get("Cargo") or "").strip()
    if not all([doc, nombre, cargo]):
        return jsonify({"success": False, "mensaje": "Cédula, nombre y cargo son requeridos"}), 400
    if Persona.query.filter_by(Documento=doc).first():
        return jsonify({"success": False, "mensaje": "Esa cédula ya está registrada"}), 400
    p = Persona(Documento=doc, Nombre=nombre, Cargo=cargo,
                Salario=d.get("Salario") or None)
    db.session.add(p)
    db.session.commit()
    return jsonify({"success": True, "id": p.id})


@admin_bp.route("/api/personas/<int:id>", methods=["PUT"])
@admin_required
def api_editar_persona(id):
    p = db.session.get(Persona, id)
    if not p:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    p.Nombre  = (d.get("Nombre") or p.Nombre).strip()
    p.Cargo   = (d.get("Cargo") or p.Cargo).strip()
    p.Salario = d.get("Salario") or p.Salario
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/personas/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_persona(id):
    p = db.session.get(Persona, id)
    if not p:
        return jsonify({"success": False}), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/personas/bulk-delete", methods=["POST"])
@admin_required
def api_bulk_delete_personas():
    data = request.get_json() or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"success": False, "mensaje": "No se enviaron IDs"}), 400
    try:
        count = Persona.query.filter(Persona.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({"success": True, "eliminados": count})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "mensaje": str(exc)}), 400


@admin_bp.route("/api/personas/export")
@admin_required
def api_exportar_personas():
    import pandas as pd
    lista = Persona.query.order_by(Persona.Nombre).all()
    df = pd.DataFrame([{
        "Documento": p.Documento,
        "Nombre":    p.Nombre,
        "Cargo":     p.Cargo,
        "Salario":   p.Salario or ""
    } for p in lista])
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="personas.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/api/personas/import", methods=["POST"])
@admin_required
def api_importar_personas():
    import pandas as pd
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"success": False, "mensaje": "No se envió archivo"}), 400
    try:
        df = pd.read_excel(io.BytesIO(archivo.read()))
        # Normalizar columnas: quitar espacios y pasar a minúsculas para búsqueda
        col_map = {str(c).strip().lower(): str(c).strip() for c in df.columns}
        df.columns = [str(c).strip() for c in df.columns]
        # Aliases aceptados por columna (en minúsculas)
        def _col(aliases):
            for a in aliases:
                if a in col_map:
                    return col_map[a]
            return None
        col_doc    = _col(["documento", "cedula", "cédula", "doc", "id"])
        col_nombre = _col(["nombre", "nombre completo", "name"])
        col_cargo  = _col(["cargo", "posicion", "posición", "position"])
        col_sal    = _col(["salario", "sueldo", "salary"])
        if not col_doc or not col_nombre or not col_cargo:
            cols_found = list(df.columns)
            return jsonify({"success": False,
                "mensaje": f"No se encontraron las columnas requeridas. Columnas en el archivo: {cols_found}. "
                           f"Se esperan: Documento (o Cedula), Nombre, Cargo."}), 400
        insertados  = 0
        actualizados = 0
        for _, row in df.iterrows():
            doc    = str(row.get(col_doc,    "")).strip()
            nombre = str(row.get(col_nombre, "")).strip()
            cargo  = str(row.get(col_cargo,  "")).strip()
            if not doc or not nombre or not cargo or doc.lower() == "nan" \
               or nombre.lower() == "nan" or cargo.lower() == "nan":
                continue
            sal_raw = row.get(col_sal) if col_sal else None
            try:
                salario = float(sal_raw) if sal_raw and not pd.isna(sal_raw) else None
            except (ValueError, TypeError):
                salario = None
            existente = Persona.query.filter_by(Documento=doc).first()
            if existente:
                existente.Nombre = nombre
                existente.Cargo  = cargo
                if salario is not None:
                    existente.Salario = salario
                actualizados += 1
            else:
                db.session.add(Persona(Documento=doc, Nombre=nombre, Cargo=cargo, Salario=salario))
                insertados += 1
        db.session.commit()
        nombres_sync = _sync_nombres_desde_personas()
        return jsonify({"success": True, "insertados": insertados,
                        "actualizados": actualizados, "nombres_sync": nombres_sync})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "mensaje": str(exc)}), 400


@admin_bp.route("/api/sync-nombres", methods=["POST"])
@admin_required
def api_sync_nombres():
    try:
        count = _sync_nombres_desde_personas()
        return jsonify({"success": True, "actualizados": count})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "mensaje": str(exc)}), 400


# ============================================================
# META OPERATIVA
# ============================================================

@admin_bp.route("/metas")
@admin_required
def metas():
    lista = MetaOperativa.query.order_by(
        MetaOperativa.contrato, MetaOperativa.Tipo_cuadrilla
    ).all()
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    return render_template("admin/metas.html", metas=lista, contratos=contratos)


@admin_bp.route("/api/metas", methods=["POST"])
@admin_required
def api_crear_meta():
    d = request.get_json() or {}
    tipo     = (d.get("Tipo_cuadrilla") or "").strip()
    contrato = (d.get("contrato") or "").strip()
    meta_val = d.get("Meta_Produccion")
    if not tipo or not contrato or meta_val is None:
        return jsonify({"success": False, "mensaje": "Tipo cuadrilla, contrato y meta son requeridos"}), 400
    try:
        meta_float = float(str(meta_val).replace(".", "").replace(",", "."))
    except (ValueError, TypeError):
        return jsonify({"success": False, "mensaje": "Valor de meta inválido"}), 400
    m = MetaOperativa(
        Tipo_cuadrilla=tipo,
        Proceso=(d.get("Proceso") or "").strip(),
        Meta_Produccion=meta_float,
        contrato=contrato
    )
    db.session.add(m)
    db.session.commit()
    return jsonify({"success": True, "id": m.id})


@admin_bp.route("/api/metas/<int:id>", methods=["PUT"])
@admin_required
def api_editar_meta(id):
    m = db.session.get(MetaOperativa, id)
    if not m:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    if d.get("Tipo_cuadrilla"):
        m.Tipo_cuadrilla = d["Tipo_cuadrilla"].strip()
    if d.get("Proceso") is not None:
        m.Proceso = d["Proceso"].strip()
    if d.get("Meta_Produccion") is not None:
        try:
            m.Meta_Produccion = float(
                str(d["Meta_Produccion"]).replace(".", "").replace(",", ".")
            )
        except (ValueError, TypeError):
            return jsonify({"success": False, "mensaje": "Valor de meta inválido"}), 400
    if d.get("contrato"):
        m.contrato = d["contrato"].strip()
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/metas/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_meta(id):
    m = db.session.get(MetaOperativa, id)
    if not m:
        return jsonify({"success": False}), 404
    db.session.delete(m)
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/metas/bulk-delete", methods=["DELETE"])
@admin_required
def api_eliminar_metas_bulk():
    ids = (request.get_json() or {}).get("ids", [])
    if not ids:
        return jsonify({"success": False, "mensaje": "Sin IDs"}), 400
    MetaOperativa.query.filter(MetaOperativa.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"success": True, "eliminados": len(ids)})


@admin_bp.route("/api/metas/export")
@admin_required
def api_exportar_metas():
    import pandas as pd
    lista = MetaOperativa.query.order_by(MetaOperativa.contrato, MetaOperativa.Tipo_cuadrilla).all()
    df = pd.DataFrame([{
        "contrato":        m.contrato,
        "Tipo_cuadrilla":  m.Tipo_cuadrilla,
        "Proceso":         m.Proceso or "",
        "Meta_Produccion": m.Meta_Produccion
    } for m in lista])
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="metas_operativas.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/api/metas/import", methods=["POST"])
@admin_required
def api_importar_metas():
    import pandas as pd
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"success": False, "mensaje": "No se envió archivo"}), 400
    try:
        df = pd.read_excel(io.BytesIO(archivo.read()))
        df.columns = [str(c).strip() for c in df.columns]
        insertados  = 0
        actualizados = 0
        for _, row in df.iterrows():
            contrato = str(row.get("contrato",        "")).strip()
            tipo     = str(row.get("Tipo_cuadrilla",  "")).strip()
            meta_raw = row.get("Meta_Produccion")
            if not contrato or not tipo or contrato == "nan" or tipo == "nan":
                continue
            if meta_raw is None or (isinstance(meta_raw, float) and pd.isna(meta_raw)):
                continue
            try:
                meta_float = float(meta_raw)
            except (ValueError, TypeError):
                continue
            proc_raw = row.get("Proceso", "")
            proceso  = str(proc_raw).strip() if proc_raw and not (isinstance(proc_raw, float) and pd.isna(proc_raw)) else ""
            existing = MetaOperativa.query.filter_by(contrato=contrato, Tipo_cuadrilla=tipo).first()
            if existing:
                existing.Meta_Produccion = meta_float
                existing.Proceso         = proceso
                actualizados += 1
            else:
                db.session.add(MetaOperativa(
                    contrato=contrato, Tipo_cuadrilla=tipo,
                    Meta_Produccion=meta_float, Proceso=proceso
                ))
                insertados += 1
        db.session.commit()
        return jsonify({"success": True, "insertados": insertados, "actualizados": actualizados})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "mensaje": str(exc)}), 400


# ============================================================
# TIPOS DE DESVÍO
# ============================================================

@admin_bp.route("/desvios")
@admin_required
def desvios():
    lista = TipoDesvio.query.order_by(TipoDesvio.tipo_desvio).all()
    return render_template("admin/desvios.html", desvios=lista)


@admin_bp.route("/api/desvios", methods=["POST"])
@admin_required
def api_crear_desvio():
    d = request.get_json() or {}
    nombre = (d.get("tipo_desvio") or "").strip()
    if not nombre:
        return jsonify({"success": False, "mensaje": "El nombre es requerido"}), 400
    if TipoDesvio.query.filter_by(tipo_desvio=nombre).first():
        return jsonify({"success": False, "mensaje": "Ya existe ese tipo de desvío"}), 400
    t = TipoDesvio(tipo_desvio=nombre)
    db.session.add(t)
    db.session.commit()
    return jsonify({"success": True, "id": t.id})


@admin_bp.route("/api/desvios/<int:id>", methods=["PUT"])
@admin_required
def api_editar_desvio(id):
    t = db.session.get(TipoDesvio, id)
    if not t:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    t.tipo_desvio = (d.get("tipo_desvio") or t.tipo_desvio).strip()
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/desvios/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_desvio(id):
    t = db.session.get(TipoDesvio, id)
    if not t:
        return jsonify({"success": False}), 404
    db.session.delete(t)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# PARÁMETROS NEO
# ============================================================

@admin_bp.route("/parametros")
@admin_required
def parametros():
    lista = ParametroNeo.query.order_by(ParametroNeo.parametroNeo).all()
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    return render_template("admin/parametros.html", parametros=lista, contratos=contratos)


@admin_bp.route("/api/parametros", methods=["POST"])
@admin_required
def api_crear_parametro():
    d = request.get_json() or {}
    nombre = (d.get("parametroNeo") or "").strip()
    if not nombre:
        return jsonify({"success": False, "mensaje": "El nombre es requerido"}), 400
    if ParametroNeo.query.filter_by(parametroNeo=nombre).first():
        return jsonify({"success": False, "mensaje": "Ya existe ese parámetro"}), 400
    p = ParametroNeo(parametroNeo=nombre)
    db.session.add(p)
    db.session.commit()
    return jsonify({"success": True, "id": p.id})


@admin_bp.route("/api/parametros/<int:id>", methods=["PUT"])
@admin_required
def api_editar_parametro(id):
    p = db.session.get(ParametroNeo, id)
    if not p:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    p.parametroNeo = (d.get("parametroNeo") or p.parametroNeo).strip()
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/parametros/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_parametro(id):
    p = db.session.get(ParametroNeo, id)
    if not p:
        return jsonify({"success": False}), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# ACTIVIDADES
# ============================================================

@admin_bp.route("/actividades")
@admin_required
def actividades():
    lista = Actividad.query.order_by(
        Actividad.contrato, Actividad.actividad
    ).all()
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    return render_template("admin/actividades.html",
                           actividades=lista, contratos=contratos)


@admin_bp.route("/api/actividades", methods=["POST"])
@admin_required
def api_crear_actividad():
    d = request.get_json() or {}
    nombre   = (d.get("actividad") or "").strip()
    contrato = (d.get("contrato") or "").strip()
    if not nombre:
        return jsonify({"success": False, "mensaje": "La actividad es requerida"}), 400
    a = Actividad(actividad=nombre, contrato=contrato or None)
    db.session.add(a)
    db.session.commit()
    return jsonify({"success": True, "id": a.id})


@admin_bp.route("/api/actividades/<int:id>", methods=["PUT"])
@admin_required
def api_editar_actividad(id):
    a = db.session.get(Actividad, id)
    if not a:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    a.actividad = (d.get("actividad") or a.actividad).strip()
    a.contrato  = d.get("contrato") or None
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/actividades/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_actividad(id):
    a = db.session.get(Actividad, id)
    if not a:
        return jsonify({"success": False}), 404
    db.session.delete(a)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# RECURSOS NEO PREDEFINIDOS (Centro Técnico / SEDE)
# ============================================================

@admin_bp.route("/recursos-neo")
@admin_required
def recursos_neo():
    lista = RecursoContrato.query.filter(
        db.or_(
            RecursoContrato.recurso.like("Centro Técnico%"),
            RecursoContrato.recurso.like("SEDE %")
        )
    ).order_by(RecursoContrato.contrato, RecursoContrato.recurso).all()
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    return render_template("admin/recursos_neo.html",
                           recursos=lista, contratos=contratos)


@admin_bp.route("/api/recursos-neo", methods=["POST"])
@admin_required
def api_crear_recurso_neo():
    d = request.get_json() or {}
    recurso  = (d.get("recurso") or "").strip()
    contrato = (d.get("contrato") or "").strip()
    if not recurso or not contrato:
        return jsonify({"success": False, "mensaje": "Recurso y contrato son requeridos"}), 400
    if RecursoContrato.query.filter_by(recurso=recurso, contrato=contrato).first():
        return jsonify({"success": False, "mensaje": "Esa combinación ya existe"}), 400
    rc = RecursoContrato(recurso=recurso, contrato=contrato)
    db.session.add(rc)
    db.session.commit()
    return jsonify({"success": True, "id": rc.id})


@admin_bp.route("/api/recursos-neo/<int:id>", methods=["PUT"])
@admin_required
def api_editar_recurso_neo(id):
    rc = db.session.get(RecursoContrato, id)
    if not rc:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    recurso  = (d.get("recurso") or "").strip()
    contrato = (d.get("contrato") or "").strip()
    if not recurso or not contrato:
        return jsonify({"success": False, "mensaje": "Recurso y contrato son requeridos"}), 400
    dup = RecursoContrato.query.filter_by(recurso=recurso, contrato=contrato).first()
    if dup and dup.id != id:
        return jsonify({"success": False, "mensaje": "Esa combinación ya existe"}), 400
    rc.recurso  = recurso
    rc.contrato = contrato
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/recursos-neo/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_recurso_neo(id):
    rc = db.session.get(RecursoContrato, id)
    if not rc:
        return jsonify({"success": False}), 404
    db.session.delete(rc)
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/recursos-neo/seed", methods=["POST"])
@admin_required
def api_seed_recursos_neo():
    from app.routes.neo import RECURSOS_EXTRA_POR_CONTRATO
    insertados = 0
    for contrato, recursos in RECURSOS_EXTRA_POR_CONTRATO.items():
        for recurso in recursos:
            if not RecursoContrato.query.filter_by(
                recurso=recurso, contrato=contrato
            ).first():
                db.session.add(RecursoContrato(recurso=recurso, contrato=contrato))
                insertados += 1
    db.session.commit()
    return jsonify({"success": True, "insertados": insertados})


# ============================================================
# CONTRATOS ASIGNADOS POR USUARIO (multi-contrato NEO)
# ============================================================

# ============================================================
# ACCIONES A TOMAR (Coordinador)
# ============================================================

_ACCIONES_SEED = [
    "Sin acción requerida",
    "Llamado de atención verbal",
    "Llamado de atención escrito",
    "Capacitación requerida",
    "Seguimiento semanal",
    "Correctivo inmediato",
]

@admin_bp.route("/acciones")
@admin_required
def acciones():
    lista = AccionTomar.query.order_by(AccionTomar.accion).all()
    return render_template("admin/acciones.html", acciones=lista)


@admin_bp.route("/api/acciones/seed", methods=["POST"])
@admin_required
def api_seed_acciones():
    insertados = 0
    for texto in _ACCIONES_SEED:
        if not AccionTomar.query.filter_by(accion=texto).first():
            db.session.add(AccionTomar(accion=texto))
            insertados += 1
    db.session.commit()
    return jsonify({"success": True, "insertados": insertados})


@admin_bp.route("/api/acciones", methods=["POST"])
@admin_required
def api_crear_accion():
    d = request.get_json() or {}
    accion = (d.get("accion") or "").strip()
    if not accion:
        return jsonify({"success": False, "mensaje": "La acción es requerida"}), 400
    if AccionTomar.query.filter_by(accion=accion).first():
        return jsonify({"success": False, "mensaje": "Esa acción ya existe"}), 409
    db.session.add(AccionTomar(accion=accion))
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/acciones/<int:id>", methods=["PUT"])
@admin_required
def api_editar_accion(id):
    a = db.session.get(AccionTomar, id)
    if not a:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    accion = (d.get("accion") or "").strip()
    if not accion:
        return jsonify({"success": False, "mensaje": "La acción es requerida"}), 400
    dup = AccionTomar.query.filter_by(accion=accion).first()
    if dup and dup.id != id:
        return jsonify({"success": False, "mensaje": "Esa acción ya existe"}), 409
    a.accion = accion
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/acciones/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_accion(id):
    a = db.session.get(AccionTomar, id)
    if not a:
        return jsonify({"success": False}), 404
    db.session.delete(a)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# PARÁMETROS COORDINADOR
# ============================================================

_PARAMETROS_COOR_SEED = [
    "Operativo", "Seguridad", "Calidad de servicio",
    "Eficiencia", "Conducta", "Otro",
]

@admin_bp.route("/parametros-coordinador")
@admin_required
def parametros_coordinador():
    lista = ParametroCoor.query.order_by(ParametroCoor.parametro).all()
    return render_template("admin/parametros_coor.html", parametros=lista)


@admin_bp.route("/api/parametros-coordinador/seed", methods=["POST"])
@admin_required
def api_seed_parametros_coor():
    insertados = 0
    for texto in _PARAMETROS_COOR_SEED:
        if not ParametroCoor.query.filter_by(parametro=texto).first():
            db.session.add(ParametroCoor(parametro=texto))
            insertados += 1
    db.session.commit()
    return jsonify({"success": True, "insertados": insertados})


@admin_bp.route("/api/parametros-coordinador", methods=["POST"])
@admin_required
def api_crear_parametro_coor():
    d = request.get_json() or {}
    parametro = (d.get("parametro") or "").strip()
    if not parametro:
        return jsonify({"success": False, "mensaje": "El parámetro es requerido"}), 400
    if ParametroCoor.query.filter_by(parametro=parametro).first():
        return jsonify({"success": False, "mensaje": "Ese parámetro ya existe"}), 409
    db.session.add(ParametroCoor(parametro=parametro))
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# EXPORTACIONES
# ============================================================

def _wb_styles():
    """Devuelve estilos comunes para exportaciones."""
    hdr_fill  = PatternFill("solid", fgColor="006d77")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="d1d9e0")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    return hdr_fill, hdr_font, hdr_align, border


@admin_bp.route("/exportar/usuarios")
@admin_required
def exportar_usuarios():
    hdr_fill, hdr_font, hdr_align, border = _wb_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    cols = ["ID", "Usuario", "Nombre Completo", "Rol", "Contrato Principal",
            "Email", "Activo", "Contratos Asignados"]
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 28

    usuarios = User.query.order_by(User.rol, User.nombre_completo).all()
    contratos_por_usuario = {}
    for uc in UserContrato.query.all():
        contratos_por_usuario.setdefault(uc.user_id, []).append(uc.contrato)

    row_font  = Font(size=10)
    row_align = Alignment(vertical="center")
    for r, u in enumerate(usuarios, 2):
        vals = [
            u.id, u.username, u.nombre_completo, u.rol,
            u.contrato or "",
            u.email or "",
            "Sí" if u.activo else "No",
            ", ".join(contratos_por_usuario.get(u.id, [])),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = row_font; cell.alignment = row_align; cell.border = border

    widths = [6, 18, 26, 14, 30, 28, 8, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name="usuarios.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/exportar/contratos")
@admin_required
def exportar_contratos():
    hdr_fill, hdr_font, hdr_align, border = _wb_styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos"

    cols = ["Código", "Contrato", "Sede", "Proceso", "Activo"]
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 28

    contratos = Contrato.query.order_by(Contrato.contrato).all()
    row_font  = Font(size=10)
    row_align = Alignment(vertical="center")
    for r, c in enumerate(contratos, 2):
        vals = [c.codigo or "", c.contrato, c.sede or "", c.proceso or "",
                "Sí" if c.activo else "No"]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = row_font; cell.alignment = row_align; cell.border = border

    widths = [12, 50, 24, 20, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name="contratos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# REPORTES — DASHBOARD Y GESTIÓN
# ============================================================

# Columnas del Excel de importación y su campo en el modelo
_IMPORT_COLS = [
    ("fecha_reporte",           "Fecha Reporte",           "date"),
    ("contrato",                "Contrato",                 "str"),
    ("recurso",                 "Recurso",                  "str"),
    ("placa",                   "Placa",                    "str"),
    ("tipo_cuadrilla",          "Tipo Cuadrilla",           "str"),
    ("meta",                    "Meta",                     "float"),
    ("tipo_actividad",          "Tipo Actividad",           "str"),
    ("orden_trabajo",           "Orden de Trabajo",         "str"),
    ("hora_inicio",             "Hora Inicio",              "time"),
    ("hora_fin",                "Hora Fin",                 "time"),
    ("tipo_incidencia",         "Tipo Incidencia",          "str"),
    ("parametro_neo",           "Parámetro NEO",            "str"),
    ("observacion",             "Observación",              "str"),
    ("duracion",                "Duración",                 "str"),
    ("impacto",                 "Impacto",                  "str"),
    ("horas_afectadas",         "Horas Afectadas",          "float"),
    ("afectacion_economica",    "Afectación Económica",     "float"),
    ("reportado_por",           "Reportado Por",            "str"),
    ("estado",                  "Estado",                   "str"),
    ("respuesta",               "Respuesta Coordinador",    "str"),
    ("parametro_coordinador",   "Parámetro Coordinador",    "str"),
    ("estado_conformidad",      "Estado Conformidad",       "str"),
    ("accion_a_tomar",          "Acción a Tomar",           "str"),
    ("respondido_por",          "Respondido Por",           "str"),
    ("fecha_respuesta",         "Fecha Respuesta",          "datetime"),
    ("conformidad_neo",         "Conformidad NEO",          "str"),
    ("observacion_conformidad", "Observación Conformidad",  "str"),
]


@admin_bp.route("/reportes")
@admin_required
def reportes():
    total = ReporteOperacional.query.count()
    return render_template("admin/reportes.html", total=total)


@admin_bp.route("/api/reportes")
@admin_required
def api_reportes():
    """Devuelve todos los reportes como JSON para Tabulator."""
    items = ReporteOperacional.query.order_by(
        ReporteOperacional.fecha_reporte.desc(),
        ReporteOperacional.id.desc()
    ).all()

    def fmt_date(d):
        return d.strftime("%d/%m/%Y") if d else ""
    def fmt_time(t):
        return t.strftime("%H:%M") if t else ""
    def fmt_dt(dt):
        return dt.strftime("%d/%m/%Y %H:%M") if dt else ""

    rows = []
    for r in items:
        rows.append({
            "id":                    r.id,
            "fecha_reporte":         fmt_date(r.fecha_reporte),
            "contrato":              r.contrato or "",
            "recurso":               r.recurso or "",
            "placa":                 r.placa or "",
            "tipo_cuadrilla":        r.tipo_cuadrilla or "",
            "meta":                  r.meta,
            "tipo_actividad":        r.tipo_actividad or "",
            "orden_trabajo":         r.orden_trabajo or "",
            "hora_inicio":           fmt_time(r.hora_inicio),
            "hora_fin":              fmt_time(r.hora_fin),
            "duracion":              r.duracion or "",
            "tipo_incidencia":       r.tipo_incidencia or "",
            "parametro_neo":         r.parametro_neo or "",
            "observacion":           r.observacion or "",
            "impacto":               r.impacto or "",
            "horas_afectadas":       r.horas_afectadas,
            "afectacion_economica":  r.afectacion_economica,
            "reportado_por":         r.reportado_por or "",
            "estado":                r.estado or "Abierto",
            "respuesta":             r.respuesta or "",
            "parametro_coordinador": r.parametro_coordinador or "",
            "estado_conformidad":    r.estado_conformidad or "",
            "accion_a_tomar":        r.accion_a_tomar or "",
            "respondido_por":        r.respondido_por or "",
            "fecha_respuesta":       fmt_dt(r.fecha_respuesta),
            "conformidad_neo":       r.conformidad_neo or "",
            "observacion_conformidad": r.observacion_conformidad or "",
        })
    return jsonify(rows)


@admin_bp.route("/reportes/plantilla")
@admin_required
def reportes_plantilla():
    """Descarga un Excel vacío con los encabezados correctos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reportes"

    header_fill  = PatternFill("solid", fgColor="006d77")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="cccccc")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (_, label, tipo) in enumerate(_IMPORT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 4, 18)

    ws.row_dimensions[1].height = 32

    # Fila de ejemplo
    ejemplo = [
        "2024-01-15", "Tolima Mantenimiento (2258)", "Cuadrilla 01", "ABC-123",
        "Cuadrilla Mantenimiento", 1.0, "Mantenimiento Correctivo", "OT-001",
        "08:00", "09:30", "Incidencia Tipo 1", "Parámetro NEO 1",
        "Observación de ejemplo", "01:30", "Medio", 1.5, 250000.0,
        "Juan Pérez", "Abierto", "", "", "", "", "", "", "", "",
    ]
    ex_font  = Font(italic=True, color="888888")
    ex_fill  = PatternFill("solid", fgColor="f0f8ff")
    for col_idx, val in enumerate(ejemplo, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font   = ex_font
        cell.fill   = ex_fill
        cell.border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_importacion_reportes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# SUPERVISORES
# ============================================================

_SUPERVISORES_SEED = [
    "ALVAREZ BUSTAMANTE PEDRO ANTONIO", "ARANAGA QUINTERO JULIAN ANDRES",
    "BOLANOS PAZ PEDRO LUIS", "ARAUJO CALERO JAIME ANDRES",
    "MARULANDA ROMERO BRAHIAM DANIEL", "OSPINA CORTES NICOLAS",
    "TORRES RUBIO YERLEY", "CANIZALES RAMIREZ JULIAN ANDRES",
    "CARMONA SANCHEZ JONNIER ALBERTO", "CASTILLO MEDINA LUISA FERNANDA",
    "LOSADA REINOSO DANIEL ALEJANDRO", "RODRIGUEZ GODOY ANDREA",
    "SALAZAR MENDOZA SAMANTA", "DELGADO RODRIGUEZ ANDRES CAMILO",
    "DONOSO PAVA CAMILO ANDRES", "ECHEVERRY LEAL EDER",
    "ASCENCIO LEYTON DIEGO ALEJANDRO", "FABIAN ROGELES ARIAS",
    "FRANCO AVALO MICHAEL CRISTHIAN", "GALEANO FRANCO ALEJANDRO",
    "BARENO FERRO BRAYAN JOAQUIN", "GUSTAVO ADOLFO FORERO GONZALEZ",
    "GUZMAN WHILINTONG", "HERNANDEZ ORTIZ FRANCISCO JAVIER",
    "HERNANDEZ TUTA HAMILTON RICARDO", "LEAL DUCUARA SANIL",
    "LLANOS LASSO DIEGO RAUL", "LOPEZ MORALES CRISTHIAN ANDRES",
    "CASTRO SOTELO YUDY PAOLA", "MARIN ESTRADA FERNAN",
    "CONDE CARRENO CRISTIAN CAMILO", "MAZUERA CORREA CAMILO",
    "MENDEZ CARTAGENA JESSIKA PAOLA", "MENDEZ HERNANDEZ VICTOR AUGUSTO",
    "MOLINA CASTRO JULIAN ANDRES", "MORALES BURITICA PEDRO ALEJANDRO",
    "NORENA MANZANO STHEFANY", "CONDE CERQUERA DIEGO FERNANDO",
    "ORTIZ ORTEGA EUDIS JESUS", "ORTIZ RIVAS NEDERLAN",
    "OSPINA REINOSA OSCAR DANILO", "ESCOBAR BARCO BRIGGIT",
    "OVALLE SANTOS JHON EDINSON FERNANDO", "PALMA GOMEZ YEISSON ANDRES",
    "GOMEZ HERRERA CLAUDIA JAZMIN", "RETAMOZO MAGREGO DANNA VANESSA",
    "NUNEZ NUSTES LUIS MIGUEL", "RODRIGUEZ GALVIS HUGO ARMANDO",
    "RINCON ALAPE JOSE ALIRIO", "ROLON REDONDO JORGE HUMBERTO",
    "ROMERO LIZARRALDE DIEGO ARMANDO", "ROPERO CANIZARES DIEGO ARMANDO",
    "SAENZ PRIETO CARLOS ARMANDO", "AMORTEGUI AGUIRRE WILLIAM STID",
    "SALDANA OREJUELA CARMEN INES", "SANCHEZ MOYA ERIKA PATRICIA",
    "SANTANDER ORTIZ JUAN SEBASTIAN", "TORRES ARENAS CRISTOBAL",
    "TORRES ARIAS ALEXANDER", "PEDROZA ANAZCO LAURA LIZETH",
    "TOVAR MIRA LEONARDO", "VANEGAS AREVALO RODRIGO ALBEIRO",
    "VIDAL CASTANEDA KAREN JULIETH", "VILLEGAS OLMOS CRISTHIAN DAVID",
    "ZAPATA ARANGO JOSE BERTULIO", "ALVARADO LEANDRO",
    "PICON AGUILAR JOHANA MARCELA", "PEREZ VASQUEZ JARED DAVID",
    "CASTRO LUIS FERNANDO",
]


@admin_bp.route("/supervisores")
@admin_required
def supervisores():
    lista = Supervisor.query.order_by(Supervisor.nombre).all()
    contratos = Contrato.query.filter_by(activo=True).order_by(Contrato.contrato).all()
    return render_template("admin/supervisores.html", supervisores=lista, contratos=contratos)


@admin_bp.route("/api/supervisores/seed", methods=["POST"])
@admin_required
def api_seed_supervisores():
    insertados = 0
    for nombre in _SUPERVISORES_SEED:
        if not Supervisor.query.filter_by(nombre=nombre).first():
            db.session.add(Supervisor(nombre=nombre))
            insertados += 1
    db.session.commit()
    return jsonify({"success": True, "insertados": insertados})


@admin_bp.route("/api/supervisores", methods=["POST"])
@admin_required
def api_crear_supervisor():
    d = request.get_json() or {}
    nombre = (d.get("nombre") or "").strip().upper()
    if not nombre:
        return jsonify({"success": False, "mensaje": "El nombre es requerido"}), 400
    if Supervisor.query.filter_by(nombre=nombre).first():
        return jsonify({"success": False, "mensaje": "Ese supervisor ya existe"}), 409
    s = Supervisor(nombre=nombre)
    ids = [int(x) for x in (d.get("contrato_ids") or []) if x]
    if ids:
        s.contratos = Contrato.query.filter(Contrato.id.in_(ids)).all()
    db.session.add(s)
    db.session.commit()
    return jsonify({"success": True, "supervisor": s.to_dict()})


@admin_bp.route("/api/supervisores/<int:id>", methods=["PUT"])
@admin_required
def api_editar_supervisor(id):
    s = db.session.get(Supervisor, id)
    if not s:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    nombre = (d.get("nombre") or "").strip().upper()
    if not nombre:
        return jsonify({"success": False, "mensaje": "El nombre es requerido"}), 400
    dup = Supervisor.query.filter_by(nombre=nombre).first()
    if dup and dup.id != id:
        return jsonify({"success": False, "mensaje": "Ese nombre ya existe"}), 409
    s.nombre = nombre
    s.activo = bool(d.get("activo", True))
    ids = [int(x) for x in (d.get("contrato_ids") or []) if x]
    s.contratos = Contrato.query.filter(Contrato.id.in_(ids)).all() if ids else []
    db.session.commit()
    return jsonify({"success": True, "supervisor": s.to_dict()})


@admin_bp.route("/api/supervisores/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_supervisor(id):
    s = db.session.get(Supervisor, id)
    if not s:
        return jsonify({"success": False}), 404
    db.session.delete(s)
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/reportes/importar", methods=["POST"])
@admin_required
def reportes_importar():
    """Importa reportes desde un archivo Excel."""
    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename.endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "mensaje": "Sube un archivo .xlsx válido"}), 400

    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"success": False, "mensaje": f"No se pudo leer el archivo: {e}"}), 400

    if len(rows) < 2:
        return jsonify({"success": False, "mensaje": "El archivo no tiene datos"}), 400

    # Mapear encabezados por posición
    header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
    label_to_field = {label: field for field, label, _ in _IMPORT_COLS}
    tipo_map       = {field: tipo for field, _, tipo in _IMPORT_COLS}

    col_map = {}  # col_index -> field_name
    for idx, h in enumerate(header_row):
        if h in label_to_field:
            col_map[idx] = label_to_field[h]

    if not col_map:
        return jsonify({"success": False, "mensaje": "No se reconocieron columnas. ¿Descargaste la plantilla?"}), 400

    def _parse(val, tipo):
        if val is None or str(val).strip() == "":
            return None
        s = str(val).strip()
        if tipo == "str":
            return s
        if tipo == "float":
            try:
                return float(str(val).replace(",", "."))
            except Exception:
                return None
        if tipo == "date":
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").date()
            except Exception:
                try:
                    return datetime.strptime(s[:10], "%d/%m/%Y").date()
                except Exception:
                    return None
        if tipo == "datetime":
            if isinstance(val, datetime):
                return val
            try:
                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except Exception:
                return None
        if tipo == "time":
            if isinstance(val, time):
                return val
            if isinstance(val, datetime):
                return val.time()
            try:
                parts = s.split(":")
                return time(int(parts[0]), int(parts[1]))
            except Exception:
                return None
        return s

    insertados = 0
    errores    = []
    for row_num, row in enumerate(rows[1:], start=2):
        try:
            kwargs = {}
            for col_idx, field in col_map.items():
                raw  = row[col_idx] if col_idx < len(row) else None
                kwargs[field] = _parse(raw, tipo_map[field])

            # Campos obligatorios mínimos
            if not kwargs.get("fecha_reporte") or not kwargs.get("contrato") \
                    or not kwargs.get("recurso") or not kwargs.get("tipo_incidencia") \
                    or not kwargs.get("parametro_neo"):
                errores.append(f"Fila {row_num}: faltan campos obligatorios (omitida)")
                continue

            if not kwargs.get("hora_inicio"):
                kwargs["hora_inicio"] = time(0, 0)
            if not kwargs.get("hora_fin"):
                kwargs["hora_fin"] = time(0, 0)

            db.session.add(ReporteOperacional(**kwargs))
            insertados += 1
        except Exception as e:
            errores.append(f"Fila {row_num}: {e}")

    db.session.commit()
    return jsonify({
        "success":    True,
        "insertados": insertados,
        "errores":    errores[:20],
    })


@admin_bp.route("/api/parametros-coordinador/<int:id>", methods=["PUT"])
@admin_required
def api_editar_parametro_coor(id):
    p = db.session.get(ParametroCoor, id)
    if not p:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    parametro = (d.get("parametro") or "").strip()
    if not parametro:
        return jsonify({"success": False, "mensaje": "El parámetro es requerido"}), 400
    dup = ParametroCoor.query.filter_by(parametro=parametro).first()
    if dup and dup.id != id:
        return jsonify({"success": False, "mensaje": "Ese parámetro ya existe"}), 409
    p.parametro = parametro
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/parametros-coordinador/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_parametro_coor(id):
    p = db.session.get(ParametroCoor, id)
    if not p:
        return jsonify({"success": False}), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================
# PLACAS POR CONTRATO (Coordinador)
# ============================================================

@admin_bp.route("/placas")
@admin_required
def placas():
    lista = PlacaContrato.query.order_by(PlacaContrato.contrato, PlacaContrato.placa).all()
    contratos = Contrato.query.filter_by(activo=True).order_by(Contrato.contrato).all()
    return render_template("admin/placas.html", placas=lista, contratos=contratos)


@admin_bp.route("/api/placas", methods=["POST"])
@admin_required
def api_crear_placa():
    d = request.get_json() or {}
    placa    = (d.get("placa") or "").strip().upper()
    contrato = (d.get("contrato") or "").strip()
    if not placa:
        return jsonify({"success": False, "mensaje": "La placa es requerida"}), 400
    if not contrato:
        return jsonify({"success": False, "mensaje": "El contrato es requerido"}), 400
    if PlacaContrato.query.filter_by(placa=placa, contrato=contrato).first():
        return jsonify({"success": False, "mensaje": "Esa placa ya existe para este contrato"}), 409
    db.session.add(PlacaContrato(placa=placa, contrato=contrato))
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/placas/<int:id>", methods=["PUT"])
@admin_required
def api_editar_placa(id):
    p = db.session.get(PlacaContrato, id)
    if not p:
        return jsonify({"success": False}), 404
    d = request.get_json() or {}
    placa    = (d.get("placa") or "").strip().upper()
    contrato = (d.get("contrato") or "").strip()
    if not placa or not contrato:
        return jsonify({"success": False, "mensaje": "Placa y contrato son requeridos"}), 400
    dup = PlacaContrato.query.filter_by(placa=placa, contrato=contrato).first()
    if dup and dup.id != id:
        return jsonify({"success": False, "mensaje": "Esa placa ya existe para este contrato"}), 409
    p.placa    = placa
    p.contrato = contrato
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/placas/<int:id>", methods=["DELETE"])
@admin_required
def api_eliminar_placa(id):
    p = db.session.get(PlacaContrato, id)
    if not p:
        return jsonify({"success": False}), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify({"success": True})


# ============================================================

@admin_bp.route("/api/usuarios/<int:id>/password", methods=["PUT"])
@admin_required
def api_cambiar_password_usuario(id):
    u = db.session.get(User, id)
    if not u:
        return jsonify({"success": False, "mensaje": "Usuario no encontrado"}), 404
    d = request.get_json() or {}
    nueva = (d.get("password") or "").strip()
    if not nueva:
        return jsonify({"success": False, "mensaje": "La contraseña no puede estar vacía"}), 400
    if len(nueva) < 4:
        return jsonify({"success": False, "mensaje": "Mínimo 4 caracteres"}), 400
    u.password_hash = nueva
    db.session.commit()
    return jsonify({"success": True})


@admin_bp.route("/api/usuarios/<int:id>/contratos", methods=["GET"])
@admin_required
def api_get_contratos_usuario(id):
    asignados = [
        uc.contrato for uc in
        UserContrato.query.filter_by(user_id=id).all()
    ]
    return jsonify({"success": True, "contratos": asignados})


@admin_bp.route("/api/usuarios/<int:id>/contratos", methods=["POST"])
@admin_required
def api_set_contratos_usuario(id):
    u = db.session.get(User, id)
    if not u:
        return jsonify({"success": False, "mensaje": "Usuario no encontrado"}), 404
    d = request.get_json() or {}
    nuevos = [c.strip() for c in (d.get("contratos") or []) if c.strip()]

    UserContrato.query.filter_by(user_id=id).delete()
    for contrato in nuevos:
        db.session.add(UserContrato(user_id=id, contrato=contrato))
    db.session.commit()
    return jsonify({"success": True, "guardados": len(nuevos)})


# ──────────────────────────────────────────────────────────────
# HE Configuración
# ──────────────────────────────────────────────────────────────

@admin_bp.route("/he/dashboard-acceso")
@admin_required
def he_dashboard_acceso():
    usuarios = User.query.filter_by(activo=True).order_by(User.nombre_completo).all()
    return render_template("admin/he_dashboard_acceso.html", usuarios=usuarios)


@admin_bp.route("/he/kpis")
@admin_required
def he_kpis():
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    cortes    = HeCorte.query.order_by(HeCorte.fecha_inicio.desc()).all()
    return render_template("admin/he_kpis.html",
                           contratos=contratos, cortes=cortes, conceptos=CONCEPTOS_HE)


@admin_bp.route("/he/cortes")
@admin_required
def he_cortes():
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    return render_template("admin/he_cortes.html", contratos=contratos)


@admin_bp.route("/he/registros")
@admin_required
def he_registros():
    contratos = Contrato.query.order_by(Contrato.contrato).all()
    cortes    = HeCorte.query.order_by(HeCorte.fecha_inicio.desc()).all()
    return render_template("admin/he_registros.html",
                           contratos=contratos, cortes=cortes, conceptos=CONCEPTOS_HE)


@admin_bp.route("/api/he/dashboard-data")
@login_required
def api_he_dashboard_data():
    if not (current_user.rol.lower() == "admin" or current_user.tiene_permiso("dashboard_he")):
        return jsonify({"error": "Sin permiso"}), 403

    # Códigos que son HE puras (cuentan para límite legal de 48h)
    CODIGOS_HE = {"03", "04", "05", "06"}

    q = HoraExtra.query
    contrato_id = request.args.get("contrato_id", type=int)
    corte_id    = request.args.get("corte_id", type=int)
    fecha_desde = request.args.get("fecha_desde", "")
    fecha_hasta = request.args.get("fecha_hasta", "")
    mes         = request.args.get("mes", "")

    if contrato_id:
        q = q.filter(HoraExtra.contrato_id == contrato_id)
    if corte_id:
        co = HeCorte.query.get(corte_id)
        if co:
            from sqlalchemy import or_ as sa_or
            q = q.filter(sa_or(HoraExtra.corte_id == corte_id,
                db.and_(HoraExtra.fecha_labor >= co.fecha_inicio, HoraExtra.fecha_labor <= co.fecha_fin)))
    if not corte_id:
        if fecha_desde:
            try: q = q.filter(HoraExtra.fecha_labor >= date.fromisoformat(fecha_desde))
            except Exception: pass
        if fecha_hasta:
            try: q = q.filter(HoraExtra.fecha_labor <= date.fromisoformat(fecha_hasta))
            except Exception: pass
        if mes and not fecha_desde and not fecha_hasta:
            try:
                yr, mo = mes.split("-")
                q = q.filter(db.extract("year", HoraExtra.fecha_labor)==int(yr),
                              db.extract("month", HoraExtra.fecha_labor)==int(mo))
            except Exception: pass

    registros = q.order_by(HoraExtra.fecha_labor).all()

    # ── KPIs ─────────────────────────────────────────────────────────────────
    total            = len(registros)
    hrs_reportadas   = sum(r.horas_reportadas or 0 for r in registros)
    hrs_autorizadas  = sum((r.horas_autorizadas or 0) for r in registros if r.horas_autorizadas is not None)
    hrs_descontadas  = sum(r.horas_autorizadas or 0 for r in registros if r.estado == "DESCONTADA")
    hrs_no_conforme  = sum(r.horas_reportadas or 0 for r in registros if r.estado == "NO CONFORME")
    hrs_conformes    = sum(r.horas_autorizadas or 0 for r in registros if r.estado == "CONFORME")
    pendientes       = sum(1 for r in registros if r.estado == "PENDIENTE")

    # ── Por tipo de HE ───────────────────────────────────────────────────────
    por_tipo = {}
    for r in registros:
        k = (r.id_concepto or "").strip().zfill(2)  # normalizar a 2 dígitos: "3" → "03"
        label = CONCEPTOS_HE.get(k, k)
        if k not in por_tipo:
            por_tipo[k] = {"codigo": k, "tipo": label, "registros": 0,
                           "hrs_reportadas": 0, "hrs_autorizadas": 0}
        por_tipo[k]["registros"]      += 1
        por_tipo[k]["hrs_reportadas"] += r.horas_reportadas or 0
        por_tipo[k]["hrs_autorizadas"]+= r.horas_autorizadas or 0
    tipos_lista = sorted(por_tipo.values(), key=lambda x: -x["hrs_reportadas"])
    top_tipo = tipos_lista[0] if tipos_lista else None

    # ── Horas por día trabajado y por día de reporte ─────────────────────────
    por_dia_todos   = {}   # agrupado por fecha_labor (día en que se trabajó)
    por_dia_reporte = {}   # agrupado por fecha_reporte (día en que se subió)
    for r in registros:
        hrs = r.horas_reportadas or 0
        if r.fecha_labor:
            d = r.fecha_labor.isoformat()
            por_dia_todos[d] = por_dia_todos.get(d, 0) + hrs
        if r.fecha_reporte:
            d = r.fecha_reporte.date().isoformat()
            por_dia_reporte[d] = por_dia_reporte.get(d, 0) + hrs

    # Rellenar todos los días del rango con 0 para por_dia_todos
    from datetime import timedelta
    rango_ini = rango_fin = None
    if corte_id:
        _co = HeCorte.query.get(corte_id)
        if _co:
            rango_ini, rango_fin = _co.fecha_inicio, _co.fecha_fin
    if not rango_ini and fecha_desde:
        try: rango_ini = date.fromisoformat(fecha_desde)
        except Exception: pass
    if not rango_fin and fecha_hasta:
        try: rango_fin = date.fromisoformat(fecha_hasta)
        except Exception: pass
    if not rango_ini and mes:
        try:
            import calendar as _cal
            yr, mo = int(mes.split("-")[0]), int(mes.split("-")[1])
            rango_ini = date(yr, mo, 1)
            rango_fin = date(yr, mo, _cal.monthrange(yr, mo)[1])
        except Exception: pass
    if rango_ini and rango_fin:
        cur = rango_ini
        while cur <= rango_fin:
            por_dia_todos.setdefault(cur.isoformat(), 0)
            por_dia_reporte.setdefault(cur.isoformat(), 0)
            cur += timedelta(days=1)

    dias         = sorted(por_dia_todos.items())
    dias_reporte = sorted(por_dia_reporte.items())

    # ── Límite legal 48h (solo HE puras 03,04,05,06) ────────────────────────
    # Si hay corte activo, los registros ya están acotados al período del corte;
    # agrupamos solo por cédula para no partir horas entre meses cuando el corte
    # cruza dos meses (ej. Jul 17 – Ago 17).
    # Sin corte, agrupamos por (cédula, mes) para mostrar el acumulado mensual.
    from collections import defaultdict
    # Para el límite legal usamos solo fecha_labor dentro del período del corte
    # (evita doble conteo del filtro OR corte_id/fecha_labor del query principal)
    co_obj = HeCorte.query.get(corte_id) if corte_id else None
    he_por_persona_mes = defaultdict(lambda: {"nombre":"","contratos":set(),"horas":0,"mes":""})
    for r in registros:
        if (r.id_concepto or "").strip().zfill(2) not in CODIGOS_HE: continue
        if not r.fecha_labor: continue
        if not (r.cedula or "").strip(): continue
        if co_obj and not (co_obj.fecha_inicio <= r.fecha_labor <= co_obj.fecha_fin):
            continue  # excluir registros fuera del rango exacto del corte
        if corte_id:
            llave = (r.cedula, "")
        else:
            llave = (r.cedula, r.fecha_labor.strftime("%Y-%m"))
        # Solo contar horas autorizadas de registros conformes o descontados
        estado = (r.estado or "").strip().upper()
        if estado not in ("CONFORME", "DESCONTADA"):
            continue
        hrs = r.horas_autorizadas or 0
        if not hrs:
            continue
        he_por_persona_mes[llave]["nombre"] = r.nombre or r.cedula
        he_por_persona_mes[llave]["mes"]    = r.fecha_labor.strftime("%Y-%m")
        if r.contrato:
            he_por_persona_mes[llave]["contratos"].add(r.contrato.contrato)
        he_por_persona_mes[llave]["horas"] += hrs
    limite_legal = []
    for (ced, mes_key), info in he_por_persona_mes.items():
        if info["horas"] >= 36:  # mostrar desde 36h (alerta 75%)
            limite_legal.append({
                "cedula":   ced,
                "nombre":   info["nombre"],
                "contrato": " / ".join(sorted(info["contratos"])) if info["contratos"] else "",
                "mes":      mes_key or info["mes"],
                "horas":    info["horas"],
                "excede":   info["horas"] >= 48,
            })
    limite_legal.sort(key=lambda x: -x["horas"])

    # ── Supervisores autorizantes ─────────────────────────────────────────────
    sup_map = defaultdict(lambda: {"registros": 0, "hrs_reportadas": 0, "hrs_autorizadas": 0})
    for r in registros:
        sup = (r.autorizacion_sup or "Sin especificar").strip()
        sup_map[sup]["registros"]      += 1
        sup_map[sup]["hrs_reportadas"] += r.horas_reportadas or 0
        sup_map[sup]["hrs_autorizadas"]+= r.horas_autorizadas or 0
    supervisores = sorted(
        [{"nombre": k, **v} for k, v in sup_map.items()],
        key=lambda x: -x["hrs_autorizadas"]
    )

    return jsonify({
        "kpis": {
            "total": total, "hrs_reportadas": hrs_reportadas,
            "hrs_autorizadas": hrs_autorizadas, "hrs_descontadas": hrs_descontadas,
            "hrs_no_conforme": hrs_no_conforme, "hrs_conformes": hrs_conformes,
            "pendientes": pendientes,
        },
        "por_tipo":      tipos_lista,
        "top_tipo":      top_tipo,
        "por_dia":         [{"fecha": d, "horas": h} for d, h in dias],
        "por_dia_reporte": [{"fecha": d, "horas": h} for d, h in dias_reporte],
        "limite_legal":  limite_legal,
        "supervisores":  supervisores,
    })


@admin_bp.route("/api/he/kpis-concepto")
@admin_required
def api_he_kpis_concepto():
    q = HoraExtra.query
    contrato_id = request.args.get("contrato_id", type=int)
    corte_id    = request.args.get("corte_id", type=int)
    fecha_desde = request.args.get("fecha_desde", "")
    fecha_hasta = request.args.get("fecha_hasta", "")
    if contrato_id:
        q = q.filter(HoraExtra.contrato_id == contrato_id)
    if corte_id:
        corte_obj = HeCorte.query.get(corte_id)
        if corte_obj:
            from sqlalchemy import or_ as sa_or
            q = q.filter(sa_or(
                HoraExtra.corte_id == corte_id,
                db.and_(HoraExtra.fecha_labor >= corte_obj.fecha_inicio,
                        HoraExtra.fecha_labor <= corte_obj.fecha_fin)))
    if fecha_desde:
        try: q = q.filter(HoraExtra.fecha_labor >= date.fromisoformat(fecha_desde))
        except Exception: pass
    if fecha_hasta:
        try: q = q.filter(HoraExtra.fecha_labor <= date.fromisoformat(fecha_hasta))
        except Exception: pass
    rows = (
        q.with_entities(
            HoraExtra.id_concepto,
            HoraExtra.tipo_he,
            db.func.count(HoraExtra.id).label("cnt"),
            db.func.coalesce(db.func.sum(HoraExtra.horas_reportadas), 0).label("hrs_rep"),
            db.func.coalesce(db.func.sum(HoraExtra.horas_autorizadas), 0).label("hrs_auth"),
        )
        .group_by(HoraExtra.id_concepto, HoraExtra.tipo_he)
        .order_by(HoraExtra.id_concepto)
        .all()
    )
    return jsonify([{
        "codigo": r.id_concepto,
        "tipo":   r.tipo_he or CONCEPTOS_HE.get(r.id_concepto, r.id_concepto),
        "registros":      int(r.cnt),
        "hrs_reportadas": float(r.hrs_rep),
        "hrs_autorizadas":float(r.hrs_auth),
    } for r in rows])



@admin_bp.route("/api/he/registros/bulk-delete", methods=["POST"])
@admin_required
def api_he_bulk_delete():
    """Elimina registros HE en un rango de fecha_labor."""
    d = request.get_json() or {}
    desde = d.get("fecha_desde")
    hasta = d.get("fecha_hasta")
    if not desde or not hasta:
        return jsonify({"ok": False, "msg": "fecha_desde y fecha_hasta son requeridos"}), 400
    try:
        from datetime import date as _date
        f_desde = _date.fromisoformat(desde)
        f_hasta = _date.fromisoformat(hasta)
    except ValueError:
        return jsonify({"ok": False, "msg": "Formato de fecha inválido (YYYY-MM-DD)"}), 400
    eliminados = HoraExtra.query.filter(
        HoraExtra.fecha_labor >= f_desde,
        HoraExtra.fecha_labor <= f_hasta
    ).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True, "eliminados": eliminados})


@admin_bp.route("/api/he/registros/export")
@admin_required
def api_he_registros_export():
    import pandas as pd
    from sqlalchemy import or_ as sa_or
    q = HoraExtra.query
    contrato_id = request.args.get("contrato_id", type=int)
    corte_id    = request.args.get("corte_id", type=int)
    estado      = request.args.get("estado", "")
    fecha_desde = request.args.get("fecha_desde", "")
    fecha_hasta = request.args.get("fecha_hasta", "")
    id_concepto = request.args.get("id_concepto", "")
    if contrato_id:
        q = q.filter(HoraExtra.contrato_id == contrato_id)
    if corte_id:
        corte_obj = HeCorte.query.get(corte_id)
        if corte_obj:
            q = q.filter(sa_or(
                HoraExtra.corte_id == corte_id,
                db.and_(HoraExtra.fecha_labor >= corte_obj.fecha_inicio,
                        HoraExtra.fecha_labor <= corte_obj.fecha_fin)))
        else:
            q = q.filter(HoraExtra.corte_id == corte_id)
    if fecha_desde:
        try: q = q.filter(HoraExtra.fecha_labor >= date.fromisoformat(fecha_desde))
        except Exception: pass
    if fecha_hasta:
        try: q = q.filter(HoraExtra.fecha_labor <= date.fromisoformat(fecha_hasta))
        except Exception: pass
    if estado:
        q = q.filter(HoraExtra.estado == estado)
    if id_concepto:
        q = q.filter(HoraExtra.id_concepto == id_concepto)
    registros = q.order_by(HoraExtra.fecha_labor.desc()).all()
    rows = []
    for r in registros:
        rows.append({
            "Fecha":           r.fecha_labor.isoformat() if r.fecha_labor else "",
            "Contrato":        r.contrato.contrato if r.contrato else "",
            "Cédula":          r.cedula,
            "Nombre":          r.nombre or "",
            "Recurso":         r.recurso or "",
            "Tipo HE":         r.tipo_he or "",
            "Hrs Reportadas":  r.horas_reportadas or 0,
            "Hrs Autorizadas": r.horas_autorizadas if r.horas_autorizadas is not None else "",
            "Estado":          r.estado or "",
            "Placa":           r.placa or "",
            "Hora Inicio":     r.hora_inicio or "",
            "Hora Fin":        r.hora_fin or "",
            "Supervisor":      r.autorizacion_sup or "",
            "Justificación":   r.justificacion or "",
            "Obs. Coordinador":r.observacion or "",
            "Obs. NEO":        r.obs_neo or "",
            "Reportado por":   r.reportado_por.nombre_completo if r.reportado_por else "",
            "Fecha Reporte":   r.fecha_reporte.strftime("%Y-%m-%d %H:%M") if r.fecha_reporte else "",
            "Validado por":    r.validado_por.nombre_completo if r.validado_por else "",
            "Fecha Validación":r.fecha_validacion.strftime("%Y-%m-%d %H:%M") if r.fecha_validacion else "",
        })
    df  = pd.DataFrame(rows)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="he_registros.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/he-configuracion")
@admin_required
def he_configuracion():
    config = {
        "password_corte_cerrado": HeConfig.get("password_corte_cerrado", ""),
    }
    return render_template("admin/he_configuracion.html", config=config)


@admin_bp.route("/api/he/config", methods=["POST"])
@admin_required
def api_he_config_guardar():
    d = request.get_json() or {}
    clave  = (d.get("clave") or "").strip()
    valor  = (d.get("valor") or "")
    claves_permitidas = {"password_corte_cerrado"}
    if clave not in claves_permitidas:
        return jsonify({"success": False, "mensaje": "Clave no permitida"}), 400
    HeConfig.set(clave, valor)
    return jsonify({"success": True})
